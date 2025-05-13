# 必要なモジュールをインポートします
import openpyxl
import statsmodels.formula.api as smf
import numpy as np
import pandas as pd

# 推定結果のスタイルを変更するためのモジュール
from statsmodels.iolib.summary2 import summary_col


def add_footer(file_path, value, sheet_name=None):
    """Excelファイルの指定したシートの最終行の下に（注）を追加する関数

    Args:
        file_path (str): Excel ファイルのパス
        value (str): 注の内容
        sheet_name (str, optional): シート名。指定しない場合は最初のシートを使用。
    """
    # Excelファイル（ワークブック）を読み込む
    wb = openpyxl.load_workbook(file_path)

    # シートを取得（指定がなければ最初のシート）
    ws = wb[sheet_name] if sheet_name in wb.sheetnames else wb.worksheets[0]

    # A列の最終行の下のセルに、注の内容を追加
    ws.cell(row=ws.max_row + 1, column=1, value=value)

    # 修正を保存
    wb.save(file_path)


def get_dummies(df, columns):
    """質的変数からダミー変数を作る関数

    Args:
        df (Dataframe): 対象 df
        columns (str[]): 質的変数の列名からなるリスト

    Returns:
        Dataframe: 新しいデータフレーム
    """
    return pd.get_dummies(df, columns=columns, dtype=int)


def process_variable_name(var_name, lang="ja"):
    """変数名を処理するための関数

    Args:
        var_name (str): 変数名
        lang (str): 言語コード ("ja", "en", "zh")

    Returns:
        dict:
            "raw": 入力値,
            "core": `**2` や `対数()` などを除いた df 中の名前,
            "formula": `np.log` や `Q()` などを追加した推定用変数名,
            "label": `の対数` などを追加したラベル名,
            "type": 変数のタイプ
            "exponent": 'exponent'タイプの場合の exponent
    """
    # 言語ごとのラベル定義
    labels = {
        "ja": {
            "log_suffix": "の対数",
            "power_suffix": "の {power} 乗",
            "intersection_suffix": "×",
        },
        "en": {
            "log_suffix": " (log)",
            "power_suffix": "^{power}",
            "intersection_suffix": " × ",
        },
        "zh": {
            "log_suffix": "的对数",
            "power_suffix": "的{power}次方",
            "intersection_suffix": "×",
        },
    }

    if ":" in var_name and not var_name.startswith(":") and not var_name.endswith(":"):
        var1, var2 = var_name.split(":")
        if ":" in var1 or ":" in var2:
            raise TypeError("var1 or var2 has :")

        parsed1 = process_variable_name(var1, lang)
        parsed2 = process_variable_name(var2, lang)
        if not isinstance(parsed1, dict) or not isinstance(parsed2, dict):
            raise TypeError("parsed1 or parsed2 is not dict")

        return {
            "raw": var_name,
            "core": [parsed1["core"], parsed2["core"]],
            "formula": f"{parsed1['formula']}:{parsed2['formula']}",
            "label": f"{parsed1['label']}{labels[lang]['intersection_suffix']}{parsed2['label']}",
            "details": [parsed1, parsed2],
            "type": "intersection",
        }
    elif var_name.startswith("log(") and var_name.endswith(")"):
        variable_core = var_name[4:-1]
        return {
            "raw": var_name,
            "core": variable_core,
            "formula": f'np.log(Q("{variable_core}"))',
            "label": f"{variable_core}{labels[lang]['log_suffix']}",
            "type": "log",
        }
    elif var_name[0:-1].endswith("**"):
        variable_core, exponent = var_name.split("**")
        return {
            "raw": var_name,
            "core": variable_core,
            "formula": f'np.power(Q("{variable_core}"), {exponent})',
            "label": f"{variable_core}{labels[lang]['power_suffix'].format(power=exponent)}",
            "exponent": int(exponent),
            "type": "power",
        }
    else:
        return {
            "raw": var_name,
            "core": var_name,
            "formula": f'Q("{var_name}")',
            "label": var_name,
            "type": "ordinary",
        }


def summary(df, var_list):
    """記述統計量を出力するための関数

    Args:
        df (DataFrame): データ
        var_list (str[]): 変数名リスト

    Returns:
        DataFrame: 記述統計量
    """
    return (
        df[var_list]
        .describe()
        .round(2)
        .loc[["count", "mean", "std", "min", "max"]]
        .transpose()
        .rename(
            columns={
                "count": "観測数",
                "mean": "平均値",
                "std": "標準偏差",
                "min": "最小値",
                "max": "最大値",
            }
        )
    )


def add_if_unique(var, unique_vars):
    #  var 是否在 unique_vars 中，基于 var["raw"] 的值
    """var["raw"]をもとに var が unique_vars にあるかどうかをチェックする関数。なければ追加

    Args:
        var (dict): 変数名の処理結果
        unique_vars (list): 受け皿
    """
    if not any(item["raw"] == var["raw"] for item in unique_vars):
        unique_vars.append(var)


def regression(variables_dicts, df, decimal_places=2):
    """回帰分析を行うための関数

    Args:
        variables_dicts (list): 被説明変数と説明変数からなる辞書のリスト
            対数の場合：  `['log(X1)', 'log(X2)']`
            べき乗の場合：   `['X1', 'X2', 'X2**2', 'X2**3']`
            ダミー変数の場合： 0-1 変数の場合はそのまま `['性別']`
                            質的変数の場合は事前に `df2 = get_dummies(df1, ['調査年'])` で処理してから、`['調査年_1990年度', '調査年_1995年度']`　を使う
                            質的変数かつカテゴリーが非常に多い場合は　そのまま `['調査年', '地域']` を使う。ただし、事後的にこれらの変数の推定結果が重要でない場合はご自身で削除。
            交差項の場合：`['X1', 'X2', 'X1:X2', 'X1**2:log(X2)']`

        df (DataFrame): データ
        decimal_places (int): 小数以下何位まで保留かを表す数値

    Returns:
        tuple: (処理済みデータフレーム, 記述統計量, 回帰結果テーブル)
    """
    # 有効なキーの定義
    valid_keys = {
        "dependent": ["被説明変数", "y", "被解释变量"],
        "independent": ["説明変数", "X", "解释变量"],
    }

    # 言語ごとのラベル定義
    labels = {
        "ja": {
            "nobs": "観測数",
            "intercept": "定数項",
            "r_squared": "決定係数",
            "r_squared_adj": "自由度修正済み決定係数",
            "log_suffix": "の対数",
            "power_suffix": "の {power} 乗",
            "intersection_suffix": "×",
        },
        "en": {
            "nobs": "N",
            "intercept": "Constant",
            "r_squared": "R-squared",
            "r_squared_adj": "Adj. R-squared",
            "log_suffix": " (log)",
            "power_suffix": "^{power}",
            "intersection_suffix": " × ",
        },
        "zh": {
            "nobs": "样本数",
            "intercept": "常数项",
            "r_squared": "决定系数",
            "r_squared_adj": "调整后决定系数",
            "log_suffix": "的对数",
            "power_suffix": "的{power}次方",
            "intersection_suffix": "×",
        },
    }

    # 入力されたキーから言語を判定
    def detect_language(key):
        if key in ["被説明変数", "説明変数"]:
            return "ja"
        elif key in ["y", "X"]:
            return "en"
        elif key in ["被解释变量", "解释变量"]:
            return "zh"
        return "ja"  # デフォルトは日本語

    models = []
    model_names = []
    exp_list = []
    dep_list = []

    # 回帰式の構築
    for i, var_dict in enumerate(variables_dicts):
        # 被説明変数のキーを取得
        dep_key = next(
            (key for key in valid_keys["dependent"] if key in var_dict), None
        )
        if dep_key is None:
            raise KeyError(
                f"被説明変数のキーが見つかりません。有効なキー: {valid_keys['dependent']}"
            )

        # 説明変数のキーを取得
        exp_key = next(
            (key for key in valid_keys["independent"] if key in var_dict), None
        )
        if exp_key is None:
            raise KeyError(
                f"説明変数のキーが見つかりません。有効なキー: {valid_keys['independent']}"
            )

        # 言語の判定
        lang = detect_language(dep_key)

        dep_var = process_variable_name(var_dict[dep_key], lang)
        model_names.append(f"（{i+1}）\n{dep_var['label']}")
        formula = "{} ~ 1".format(dep_var["formula"])

        # log(Y), Y**2, Y を異なるアイテムとして識別するために、"raw" を使う
        if not any(item["raw"] == dep_var["raw"] for item in dep_list):
            dep_list.append(dep_var)

        for var in var_dict[exp_key]:
            exp_var = process_variable_name(var, lang)
            formula += " + {}".format(exp_var["formula"])

            # log(X), X**2, X を異なるアイテムとして識別するために、"raw" を使う
            if not any(item["raw"] == exp_var["raw"] for item in exp_list):
                exp_list.append(exp_var)

        model = smf.ols(formula=formula, data=df).fit(cov_type="HC1")
        models.append(model)

    # 説明変数の並び替え
    regressor_order = [var["formula"] for var in exp_list] + ["Intercept"]

    # 回帰分析
    df_result = summary_col(
        models,
        regressor_order=regressor_order,
        float_format=f"%.{decimal_places}f",
        model_names=model_names,
        info_dict={labels[lang]["nobs"]: lambda x: str(int(x.nobs))},
        stars=True,
    ).tables[0]

    # 変数名の変更
    rename_dict = {var["formula"]: var["label"] for var in exp_list}

    rename_dict.update(
        {
            "Intercept": labels[lang]["intercept"],
            "R-squared": labels[lang]["r_squared"],
            "R-squared Adj.": labels[lang]["r_squared_adj"],
        }
    )
    df_result = df_result.rename(index=rename_dict)

    # 使用した変数からなるサンプル
    unique_var_list = []
    for var in dep_list + exp_list:
        if var["type"] == "intersection":
            for sub_var in var["details"]:
                add_if_unique(sub_var, unique_var_list)
        else:
            add_if_unique(var, unique_var_list)

    # unique_var_list をもとに、サンプルを作成
    df_coppied = df.copy()
    used_var_list = []
    for var in unique_var_list:
        if var["type"] == "log":
            df_coppied[var["label"]] = np.log(df_coppied[var["core"]])
            used_var_list.append(var["label"])
        elif var["type"] == "power":
            df_coppied[var["label"]] = np.power(
                df_coppied[var["core"]], var["exponent"]
            )
            used_var_list.append(var["label"])
        elif var["type"] == "intersection":
            var1, var2 = var["core"]
            df_coppied[var["label"]] = df_coppied[var1] * df_coppied[var2]
            used_var_list.append(var["label"])
        else:
            used_var_list.append(var["label"])

    # 記述統計量
    summary_result = summary(df_coppied, used_var_list)

    return (df_coppied, summary_result, df_result)
